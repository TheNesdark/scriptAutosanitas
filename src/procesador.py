import pandas as pd
import datetime as date
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from src.validaciones import validar_telefono


NIT = "812000344"
CSBH = "11033"
NPR = "ESE HOSPITAL LOCAL DE MONTELIBANO"

def procesar_archivo(archivo_entrada, archivo_salida):
    datos = []
    ex = pd.read_excel(archivo_entrada, engine="openpyxl", header=1)

    for _, fila in ex.iterrows():

        IsAnestecia = str(fila["Descripcion_Prestacion"]).strip().upper().__contains__("ANESTESIA")
        IsSedacion = str(fila["Descripcion_Prestacion"]).strip().upper().__contains__("SEDACIÓN")
        IsContraste = str(fila["Descripcion_Prestacion"]).strip().upper().__contains__("CONTRASTE")
        IsComparativo = str(fila["Descripcion_Prestacion"]).strip().upper().__contains__("COMPARATIVO")
        IsBilateral = str(fila["Descripcion_Prestacion"]).strip().upper().__contains__("BILATERAL")



        if str(fila["Especialidad_Remitente"]).strip().upper().startswith("PYP"):
            continue

        if str(fila["Descripcion_Prestacion"]).strip().upper().__contains__("GENERAL"):
            continue

        datos.append({
            "FECHA_ENVIO": date.datetime.now().strftime("%d/%m/%Y"),
            "NIT_PRESTADOR_REMITENTE": NIT,
            "CODIGO_SUCURSAL_BH": CSBH,
            "NOMBRE_PRESTADOR_REMITENTE": NPR,
            "TIPO_IDENTIFICACION_DEL_AFILIADO": fila["Tipo_Identificación_del_Afiliado"],
            "NUMERO_DE_IDENTIFICACION_DEL_AFILIADO": str(fila["Numero_de_Identificación_del_Afiliado"]),
            "NOMBRE_PACIENTE": fila["Nombre_Paciente"],
            "TELEFONO_CELULAR_1": validar_telefono(fila["Telefono_Celular_1"]),
            "TELEFONO_CELULAR_2": validar_telefono(fila["Telefono_Celular_2"]),
            "FECHA_ATENCION": fila["Fecha_Atencion"],
            "CIE10": fila["cie10"],
            "NOMBRE_MEDICO": fila["Nombre_Medico"],
            "CODIGO_ESPECIALIDAD_REMITENTE": fila["Codigo_Especialidad_Remitente"],
            "ESPECIALIDAD_REMITENTE": fila["Especialidad_Remitente"],
            "CODIGO_CUPS_PRESTACION": fila["Codigo_CUPS_Prestacion"],
            "DESCRIPCION_PRESTACION": fila["Descripcion_Prestacion"],
            "CANTIDAD": fila["Cantidad"],
            "JUSTIFICACION_CLINICA": fila["Justificacion_Clinica"],
            "EDAD_GESTACIONAL_SEMANAS": fila["Edad_Gestacional"],
            "ANESTESIA": "x" if IsAnestecia else "",
            "SEDACION": "x" if IsSedacion else "",
            "CONTRASTE": "x" if IsContraste else "",
            "COMPARATIVO": "x" if IsComparativo else "",
            "BILATERAL": "x" if IsBilateral else "",
            "NOAPLICA": "x" if not (IsAnestecia or IsSedacion or IsContraste or IsComparativo or IsBilateral) else "",
        })

    nv = pd.DataFrame(datos)
    nv.to_excel(archivo_salida, index=False, engine="openpyxl", sheet_name="plantilla")

    # Ajustar ancho de columnas al contenido
 
    wb = load_workbook(archivo_salida)
    ws = wb.active
    for column_cells in ws.columns:
        max_length = 0
        col_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Aplicar bordes a todas las celdas
    thin_border = Border(left=Side(style="thin"),
                         right=Side(style="thin"),
                         top=Side(style="thin"),
                         bottom=Side(style="thin"))

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    wb.save(archivo_salida)

    return len(datos)
